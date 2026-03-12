"""Regression tests for ip (peak short-circuit current) calculation.

These tests verify:
1. Transformer Pkr correctly affects equivalent resistance
2. PSU correction factor is applied to impedance
3. Network topology (meshed vs radial) is detected
4. XLSX round-trip preserves Pkr and Ra
"""

import math
import pytest

from app.engine import (
    Network,
    Busbar,
    ExternalGrid,
    Line,
    Transformer2W,
    SynchronousGenerator,
    PowerStationUnit,
    NeutralGrounding,
    calculate_short_circuit,
)
from app.engine.elements import ComplexImpedance


class TestTransformerImpedanceWithPkr:
    """Test transformer impedance calculation with short-circuit losses (Pkr)."""

    def test_reference_case_transformer_impedance(self):
        """
        Verify transformer impedance for reference case:
        - Sn = 10 MVA
        - Un_hv = 22 kV
        - Un_lv = 6.3 kV
        - uk = 7.08 %
        - Pkr = 60.915 kW

        Expected (at 22 kV):
        - rk_pu = 0.0060915 pu
        - zk_pu = 0.0708 pu
        - xk_pu = sqrt(0.0708^2 - 0.0060915^2) = 0.0705375 pu
        - Zbase = 48.4 ohm
        - R = 0.29483 ohm
        - X = 3.41401 ohm
        """
        tr = Transformer2W(
            id="tr_ref",
            bus_hv="bus_22",
            bus_lv="bus_6",
            Sn=10.0,
            Un_hv=22.0,
            Un_lv=6.3,
            uk_percent=7.08,
            Pkr=60.915,
            vector_group="Dyn11",
        )

        # Get impedance referred to HV side (22 kV)
        Z1, Z2, Z0 = tr.get_impedance(22.0)

        # Expected values
        expected_R = 0.29483  # ohm
        expected_X = 3.41401  # ohm

        # Tolerance: 0.1%
        assert abs(Z1.r - expected_R) / expected_R < 0.001, \
            f"R1 = {Z1.r:.5f}, expected {expected_R:.5f}"
        assert abs(Z1.x - expected_X) / expected_X < 0.001, \
            f"X1 = {Z1.x:.5f}, expected {expected_X:.5f}"

    def test_reference_case_impedance_at_lv_side(self):
        """
        Verify transformer impedance referred to LV side (6.3 kV).

        Expected (at 6.3 kV):
        - Zbase = 3.969 ohm
        - R = 0.02418 ohm
        - X = 0.27996 ohm
        """
        tr = Transformer2W(
            id="tr_ref",
            bus_hv="bus_22",
            bus_lv="bus_6",
            Sn=10.0,
            Un_hv=22.0,
            Un_lv=6.3,
            uk_percent=7.08,
            Pkr=60.915,
            vector_group="Dyn11",
        )

        # Get impedance referred to LV side (6.3 kV)
        Z1, Z2, Z0 = tr.get_impedance(6.3)

        # Expected values
        expected_R = 0.02418  # ohm
        expected_X = 0.27996  # ohm

        # Tolerance: 0.5% (transformation ratio compounds errors)
        assert abs(Z1.r - expected_R) / expected_R < 0.005, \
            f"R1 = {Z1.r:.5f}, expected {expected_R:.5f}"
        assert abs(Z1.x - expected_X) / expected_X < 0.005, \
            f"X1 = {Z1.x:.5f}, expected {expected_X:.5f}"

    def test_zero_pkr_gives_zero_resistance(self):
        """Verify that Pkr=0 results in zero resistance."""
        tr = Transformer2W(
            id="tr_zero_r",
            bus_hv="bus_22",
            bus_lv="bus_6",
            Sn=10.0,
            Un_hv=22.0,
            Un_lv=6.3,
            uk_percent=7.08,
            Pkr=0.0,  # No losses
            vector_group="Dyn11",
        )

        Z1, _, _ = tr.get_impedance(22.0)

        assert Z1.r == 0.0, f"R1 should be 0 with Pkr=0, got {Z1.r}"
        assert Z1.x > 0, "X1 should be > 0"


class TestPkrAffectsIp:
    """Test that Pkr affects ip through R/X ratio."""

    @pytest.fixture
    def network_with_pkr(self):
        """Create network with transformer having realistic Pkr."""
        net = Network(name="Pkr Test - With Losses")

        net.add_element(Busbar(id="bus_110", Un=110.0))
        net.add_element(Busbar(id="bus_22", Un=22.0))

        net.add_element(ExternalGrid(
            id="grid",
            bus_id="bus_110",
            Sk_max=2000.0,
            Sk_min=1500.0,
            rx_ratio=0.1,
        ))

        net.add_element(Transformer2W(
            id="tr1",
            bus_hv="bus_110",
            bus_lv="bus_22",
            Sn=40.0,
            Un_hv=110.0,
            Un_lv=22.0,
            uk_percent=10.0,
            Pkr=200.0,  # 200 kW losses = significant R
            vector_group="YNyn0",
        ))

        return net

    @pytest.fixture
    def network_without_pkr(self):
        """Create network with transformer having zero Pkr."""
        net = Network(name="Pkr Test - No Losses")

        net.add_element(Busbar(id="bus_110", Un=110.0))
        net.add_element(Busbar(id="bus_22", Un=22.0))

        net.add_element(ExternalGrid(
            id="grid",
            bus_id="bus_110",
            Sk_max=2000.0,
            Sk_min=1500.0,
            rx_ratio=0.1,
        ))

        net.add_element(Transformer2W(
            id="tr1",
            bus_hv="bus_110",
            bus_lv="bus_22",
            Sn=40.0,
            Un_hv=110.0,
            Un_lv=22.0,
            uk_percent=10.0,
            Pkr=0.0,  # Zero losses = R=0
            vector_group="YNyn0",
        ))

        return net

    def test_pkr_reduces_ip(self, network_with_pkr, network_without_pkr):
        """
        Test that adding Pkr increases R/X ratio, which reduces kappa and ip.

        With Pkr=0: R/X is lower, kappa is higher, ip is higher
        With Pkr>0: R/X is higher, kappa is lower, ip is lower
        """
        # Run calculation with Pkr
        run_with = calculate_short_circuit(
            network=network_with_pkr,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max"
        )

        # Run calculation without Pkr
        run_without = calculate_short_circuit(
            network=network_without_pkr,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max"
        )

        assert run_with.is_success
        assert run_without.is_success

        result_with = run_with.results[0]
        result_without = run_without.results[0]

        # Ik3 should be nearly the same (Pkr affects R, not |Z| significantly)
        ik_diff_percent = abs(result_with.Ik - result_without.Ik) / result_without.Ik * 100
        assert ik_diff_percent < 2.0, \
            f"Ik3 should differ by <2%, got {ik_diff_percent:.2f}%"

        # R/X should be higher with Pkr
        assert result_with.R_X_ratio > result_without.R_X_ratio, \
            f"R/X with Pkr ({result_with.R_X_ratio:.4f}) should be > " \
            f"R/X without Pkr ({result_without.R_X_ratio:.4f})"

        # ip should be lower with Pkr (higher R/X = lower kappa)
        assert result_with.ip < result_without.ip, \
            f"ip with Pkr ({result_with.ip:.3f}) should be < " \
            f"ip without Pkr ({result_without.ip:.3f})"

    def test_pkr_change_has_noticeable_effect(self, network_with_pkr, network_without_pkr):
        """Test that Pkr change has noticeable effect on ip (>1%)."""
        run_with = calculate_short_circuit(
            network=network_with_pkr,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max"
        )
        run_without = calculate_short_circuit(
            network=network_without_pkr,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max"
        )

        result_with = run_with.results[0]
        result_without = run_without.results[0]

        ip_diff_percent = abs(result_with.ip - result_without.ip) / result_without.ip * 100

        # Difference should be noticeable (>1%)
        assert ip_diff_percent > 1.0, \
            f"ip difference should be >1%, got {ip_diff_percent:.2f}%"


class TestPSUCorrectionFactor:
    """Test that PSU correction factor is applied to impedance."""

    @pytest.fixture
    def psu_network(self):
        """Create network with PSU."""
        net = Network(name="PSU Test")

        net.add_element(Busbar(id="bus_110", Un=110.0))
        net.add_element(Busbar(id="bus_gen", Un=10.5))

        # Generator
        net.add_element(SynchronousGenerator(
            id="gen1",
            bus_id="bus_gen",
            Sn=100.0,
            Un=10.5,
            Xd_pp=18.0,
            Ra=0.3,
            cos_phi=0.85,
        ))

        # Block transformer
        net.add_element(Transformer2W(
            id="tr_block",
            bus_hv="bus_110",
            bus_lv="bus_gen",
            Sn=100.0,
            Un_hv=115.0,
            Un_lv=10.5,
            uk_percent=12.0,
            Pkr=400.0,
            vector_group="YNd11",
        ))

        # PSU
        net.add_element(PowerStationUnit(
            id="psu1",
            generator_id="gen1",
            transformer_id="tr_block",
            has_oltc=True,
        ))

        return net

    def test_psu_correction_factor_calculated(self, psu_network):
        """Test that PSU reports correction factor in results."""
        run = calculate_short_circuit(
            network=psu_network,
            fault_types=["Ik3"],
            fault_buses=["bus_110"],
            mode="max"
        )

        assert run.is_success
        result = run.results[0]

        # Should have KS or KSO in correction factors
        has_k_factor = any(
            k.startswith("KS") for k in result.correction_factors.keys()
        )
        assert has_k_factor, \
            f"Should have KS/KSO factor, got: {result.correction_factors.keys()}"

    def test_psu_impedance_is_modified(self):
        """Test that PSU correction factor modifies combined impedance."""
        from app.engine.psu import PowerStationUnit

        # Create generator
        gen = SynchronousGenerator(
            id="gen",
            bus_id="bus_gen",
            Sn=100.0,
            Un=10.5,
            Xd_pp=18.0,
            Ra=0.3,
            cos_phi=0.85,
        )

        # Create transformer
        tr = Transformer2W(
            id="tr",
            bus_hv="bus_110",
            bus_lv="bus_gen",
            Sn=100.0,
            Un_hv=115.0,
            Un_lv=10.5,
            uk_percent=12.0,
            Pkr=400.0,
            vector_group="YNd11",
        )

        # Create PSU
        psu = PowerStationUnit(
            id="psu",
            generator_id="gen",
            transformer_id="tr",
            has_oltc=True,
        )
        psu.set_references(gen, tr)

        # Get combined impedance at ref_voltage = 110 kV
        # This uses Un_Q = 110 kV for the KS calculation (network voltage)
        ref_voltage = 110.0
        Z1, Z2, Z0 = psu.get_combined_impedance(ref_voltage=ref_voltage, c=1.1)

        # Get correction factor with network voltage (should match what combined_impedance uses)
        factor_name, K = psu.get_correction_factor(c=1.1, Un_Q=ref_voltage)

        # Get raw gen + tr impedances
        Z1_gen, _, _ = gen.get_impedance(ref_voltage)
        Z1_tr, _, _ = tr.get_impedance(ref_voltage)
        Z1_raw = Z1_gen + Z1_tr

        # Combined impedance should be scaled by 1/K
        if K != 1.0:
            expected_r = Z1_raw.r / K
            expected_x = Z1_raw.x / K

            assert abs(Z1.r - expected_r) / expected_r < 0.001, \
                f"R should be {expected_r:.6f}, got {Z1.r:.6f}"
            assert abs(Z1.x - expected_x) / expected_x < 0.001, \
                f"X should be {expected_x:.6f}, got {Z1.x:.6f}"


class TestMeshedNetworkDetection:
    """Test meshed network detection and warning."""

    @pytest.fixture
    def radial_network(self):
        """Create simple radial network (single source)."""
        net = Network(name="Radial Network")

        net.add_element(Busbar(id="bus_110", Un=110.0))
        net.add_element(Busbar(id="bus_22", Un=22.0))

        net.add_element(ExternalGrid(
            id="grid",
            bus_id="bus_110",
            Sk_max=2000.0,
            Sk_min=1500.0,
            rx_ratio=0.1,
        ))

        net.add_element(Transformer2W(
            id="tr1",
            bus_hv="bus_110",
            bus_lv="bus_22",
            Sn=40.0,
            Un_hv=110.0,
            Un_lv=22.0,
            uk_percent=10.0,
            Pkr=150.0,
            vector_group="YNyn0",
        ))

        return net

    @pytest.fixture
    def meshed_network(self):
        """Create meshed network (two sources, loop)."""
        net = Network(name="Meshed Network")

        net.add_element(Busbar(id="bus_A", Un=110.0))
        net.add_element(Busbar(id="bus_B", Un=110.0))
        net.add_element(Busbar(id="bus_C", Un=110.0))

        # Two sources
        net.add_element(ExternalGrid(
            id="grid_A",
            bus_id="bus_A",
            Sk_max=3000.0,
            Sk_min=2500.0,
            rx_ratio=0.1,
        ))

        net.add_element(ExternalGrid(
            id="grid_B",
            bus_id="bus_B",
            Sk_max=2000.0,
            Sk_min=1500.0,
            rx_ratio=0.12,
        ))

        # Lines forming a loop
        net.add_element(Line(
            id="line_AB",
            bus_from="bus_A",
            bus_to="bus_B",
            length=50.0,
            r1_per_km=0.1,
            x1_per_km=0.4,
            r0_per_km=0.3,
            x0_per_km=1.2,
        ))

        net.add_element(Line(
            id="line_BC",
            bus_from="bus_B",
            bus_to="bus_C",
            length=30.0,
            r1_per_km=0.1,
            x1_per_km=0.4,
            r0_per_km=0.3,
            x0_per_km=1.2,
        ))

        net.add_element(Line(
            id="line_CA",
            bus_from="bus_C",
            bus_to="bus_A",
            length=40.0,
            r1_per_km=0.1,
            x1_per_km=0.4,
            r0_per_km=0.3,
            x0_per_km=1.2,
        ))

        return net

    def test_radial_no_warning(self, radial_network):
        """Test that radial network does not produce mesh warning."""
        run = calculate_short_circuit(
            network=radial_network,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max"
        )

        assert run.is_success
        result = run.results[0]

        mesh_warnings = [w for w in result.warnings if "mesh" in w.lower()]
        assert len(mesh_warnings) == 0, \
            f"Radial network should not have mesh warning: {mesh_warnings}"

    def test_meshed_has_warning(self, meshed_network):
        """Test that meshed network produces warning about kappa method."""
        run = calculate_short_circuit(
            network=meshed_network,
            fault_types=["Ik3"],
            fault_buses=["bus_C"],
            mode="max"
        )

        assert run.is_success
        result = run.results[0]

        mesh_warnings = [w for w in result.warnings if "mesh" in w.lower()]
        assert len(mesh_warnings) > 0, \
            f"Meshed network should have warning, got: {result.warnings}"


class TestXLSXRoundTrip:
    """Test that XLSX export/import preserves Pkr and Ra."""

    def test_export_includes_pkr_and_ra(self):
        """Test that export includes Pkr and Ra columns."""
        from app.services.export_xlsx import generate_calculation_report
        from app.models import CalculationRun as DBCalcRun, Project, NetworkVersion
        from unittest.mock import MagicMock

        # Create mock objects
        project = MagicMock(spec=Project)
        project.name = "Test Project"
        project.description = "Test"

        version = MagicMock(spec=NetworkVersion)
        version.version_number = 1
        version.elements = {
            "busbars": [{"id": "bus1", "name": "Bus 1", "Un": 22.0}],
            "transformers_2w": [{
                "id": "tr1",
                "name": "TR1",
                "bus_hv": "bus_hv",
                "bus_lv": "bus_lv",
                "Sn": 40.0,
                "Un_hv": 110.0,
                "Un_lv": 22.0,
                "uk_percent": 10.0,
                "Pkr": 150.0,  # This should be included
                "vector_group": "YNyn0"
            }],
            "generators": [{
                "id": "gen1",
                "name": "Gen 1",
                "bus_id": "bus1",
                "Sn": 50.0,
                "Un": 22.0,
                "Xd_pp": 15.0,
                "Ra": 0.5,  # This should be included
                "cos_phi": 0.85
            }]
        }

        run = MagicMock(spec=DBCalcRun)
        run.calculation_mode.value = "max"
        run.fault_types = ["Ik3"]
        run.completed_at = None
        run.engine_version = "1.0.0"
        run.results = []

        # Generate report
        xlsx_buffer = generate_calculation_report(run, project, version)

        # Read back with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_buffer)

        # Check "Prvky siete" sheet
        ws = wb["Prvky siete"]

        # Find transformer headers
        found_pkr = False
        found_ra = False

        for row in ws.iter_rows(values_only=True):
            if row and "Pkr" in str(row):
                found_pkr = True
            if row and "Ra" in str(row):
                found_ra = True

        assert found_pkr, "Pkr column should be in XLSX export"
        assert found_ra, "Ra column should be in XLSX export"

    def test_import_preserves_pkr(self):
        """Test that import preserves Pkr values."""
        from app.services.import_network import import_from_json
        import json

        data = {
            "busbars": [
                {"id": "bus_hv", "Un": 110.0},
                {"id": "bus_lv", "Un": 22.0}
            ],
            "transformers_2w": [{
                "id": "tr1",
                "bus_hv": "bus_hv",
                "bus_lv": "bus_lv",
                "Sn": 40.0,
                "Un_hv": 110.0,
                "Un_lv": 22.0,
                "uk_percent": 10.0,
                "Pkr": 150.0,
                "vector_group": "YNyn0"
            }]
        }

        validated = import_from_json(json.dumps(data))

        tr = validated["transformers_2w"][0]
        assert tr["Pkr"] == 150.0, f"Pkr should be 150.0, got {tr['Pkr']}"

    def test_import_preserves_ra(self):
        """Test that import preserves Ra values in % (values >= 1.0 stay as-is)."""
        from app.services.import_network import import_from_json
        import json

        data = {
            "busbars": [
                {"id": "bus1", "Un": 22.0}
            ],
            "generators": [{
                "id": "gen1",
                "bus_id": "bus1",
                "Sn": 50.0,
                "Un": 22.0,
                "Xd_pp": 15.0,
                "Ra": 1.5,  # Value >= 1.0 won't be auto-converted
                "cos_phi": 0.85
            }]
        }

        validated = import_from_json(json.dumps(data))

        gen = validated["generators"][0]
        assert gen["Ra"] == 1.5, f"Ra should be 1.5, got {gen['Ra']}"

    def test_import_converts_ra_from_pu(self):
        """Test that import auto-converts Ra from p.u. to % when value < 1.0."""
        from app.services.import_network import import_from_json
        import json

        data = {
            "busbars": [
                {"id": "bus1", "Un": 22.0}
            ],
            "generators": [{
                "id": "gen1",
                "bus_id": "bus1",
                "Sn": 50.0,
                "Un": 22.0,
                "Xd_pp": 15.0,
                "Ra": 0.005,  # 0.5% as p.u.
                "cos_phi": 0.85
            }]
        }

        validated = import_from_json(json.dumps(data))

        gen = validated["generators"][0]
        assert gen["Ra"] == 0.5, f"Ra should be 0.5% (auto-converted from 0.005 p.u.), got {gen['Ra']}"


class TestImportNormalization:
    """Test import field name normalization for alternative formats."""

    def test_normalize_transformer_field_names(self):
        """Test that alternative transformer field names are normalized."""
        from app.services.import_network import import_from_json
        import json

        # Alternative format with different field names
        data = {
            "busbars": [
                {"id": "hv", "Un": 22.0},
                {"id": "lv", "Un": 6.3}
            ],
            "transformers": [{  # 'transformers' instead of 'transformers_2w'
                "id": "tr1",
                "hv_bus_id": "hv",  # 'hv_bus_id' instead of 'bus_hv'
                "lv_bus_id": "lv",  # 'lv_bus_id' instead of 'bus_lv'
                "sn_mva": 10.0,     # 'sn_mva' instead of 'Sn'
                "un1_kv": 22.0,     # 'un1_kv' instead of 'Un_hv'
                "un2_kv": 6.3,      # 'un2_kv' instead of 'Un_lv'
                "uk_percent": 7.08,
                "pkr_kw": 60.915,   # 'pkr_kw' instead of 'Pkr'
                "vector_group": "Dyn11"
            }]
        }

        validated = import_from_json(json.dumps(data))

        assert "transformers_2w" in validated
        tr = validated["transformers_2w"][0]
        assert tr["bus_hv"] == "hv"
        assert tr["bus_lv"] == "lv"
        assert tr["Sn"] == 10.0
        assert tr["Un_hv"] == 22.0
        assert tr["Un_lv"] == 6.3
        assert tr["Pkr"] == 60.915

    def test_normalize_generator_field_names(self):
        """Test that alternative generator field names are normalized."""
        from app.services.import_network import import_from_json
        import json

        data = {
            "busbars": [
                {"id": "bus1", "Un": 6.3}
            ],
            "generators": [{
                "id": "gen1",
                "bus_id": "bus1",
                "sn_mva": 10.0,    # 'sn_mva' instead of 'Sn'
                "un_kv": 6.3,      # 'un_kv' instead of 'Un'
                "xdpp_pu": 0.15,   # 'xdpp_pu' instead of 'Xd_pp' (also p.u.)
                "cos_phi": 0.85
            }]
        }

        validated = import_from_json(json.dumps(data))

        gen = validated["generators"][0]
        assert gen["Sn"] == 10.0
        assert gen["Un"] == 6.3
        assert gen["Xd_pp"] == 15.0  # Converted from 0.15 p.u. to 15%

    def test_normalize_bus_field_names(self):
        """Test that alternative bus field names are normalized."""
        from app.services.import_network import import_from_json
        import json

        data = {
            "buses": [{  # 'buses' instead of 'busbars'
                "id": "bus1",
                "un_kv": 22.0  # 'un_kv' instead of 'Un'
            }]
        }

        validated = import_from_json(json.dumps(data))

        assert "busbars" in validated
        bus = validated["busbars"][0]
        assert bus["Un"] == 22.0

    def test_normalize_generator_ra_from_ohms(self):
        """Test conversion of Ra from ohms to percent."""
        from app.services.import_network import import_from_json
        import json

        # Generator with 10 MVA, 6.3 kV
        # Zbase = 6.3^2 / 10 = 3.969 Ω
        # Ra_ohm = 0.01 Ω
        # Ra_pu = 0.01 / 3.969 = 0.00252
        # Ra_% = 0.252
        data = {
            "busbars": [
                {"id": "bus1", "Un": 6.3}
            ],
            "generators": [{
                "id": "gen1",
                "bus_id": "bus1",
                "sn_mva": 10.0,
                "un_kv": 6.3,
                "xdpp_pu": 0.15,
                "ra_ohm": 0.01,  # Ra in ohms
                "cos_phi": 0.85
            }]
        }

        validated = import_from_json(json.dumps(data))

        gen = validated["generators"][0]
        # Ra should be converted to % via p.u.
        expected_ra = (0.01 / 3.969) * 100  # ≈ 0.252%
        assert abs(gen["Ra"] - expected_ra) < 0.01


if __name__ == "__main__":
    pytest.main([__file__, "-v"])


class TestGeneratorContributionToShortCircuit:
    """Verify synchronous generator contribution is included in short-circuit results."""

    def _build_network(self, generator_in_service: bool):
        net = Network(name="Generator Contribution Toggle")

        net.add_element(Busbar(id="bus_110", Un=110.0))
        net.add_element(Busbar(id="bus_22", Un=22.0))
        net.add_element(Busbar(id="bus_gen", Un=10.5))

        net.add_element(ExternalGrid(
            id="grid",
            bus_id="bus_110",
            Sk_max=2000.0,
            Sk_min=1500.0,
            rx_ratio=0.1,
        ))

        net.add_element(Transformer2W(
            id="tr_main",
            bus_hv="bus_110",
            bus_lv="bus_22",
            Sn=40.0,
            Un_hv=110.0,
            Un_lv=22.0,
            uk_percent=10.0,
            Pkr=150.0,
            vector_group="YNyn0",
            neutral_grounding_hv=NeutralGrounding.GROUNDED,
            neutral_grounding_lv=NeutralGrounding.GROUNDED,
        ))

        net.add_element(Transformer2W(
            id="tr_block",
            bus_hv="bus_22",
            bus_lv="bus_gen",
            Sn=20.0,
            Un_hv=22.0,
            Un_lv=10.5,
            uk_percent=12.0,
            Pkr=100.0,
            vector_group="YNd11",
            neutral_grounding_hv=NeutralGrounding.GROUNDED,
            neutral_grounding_lv=NeutralGrounding.ISOLATED,
        ))

        net.add_element(SynchronousGenerator(
            id="g1",
            bus_id="bus_gen",
            Sn=20.0,
            Un=10.5,
            Xd_pp=15.0,
            Ra=0.2,
            cos_phi=0.85,
            in_service=generator_in_service,
        ))

        return net

    def test_generator_in_service_increases_ik3_and_ip(self):
        net_off = self._build_network(generator_in_service=False)
        net_on = self._build_network(generator_in_service=True)

        run_off = calculate_short_circuit(
            network=net_off,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max",
        )
        run_on = calculate_short_circuit(
            network=net_on,
            fault_types=["Ik3"],
            fault_buses=["bus_22"],
            mode="max",
        )

        assert run_off.is_success and run_on.is_success

        res_off = run_off.results[0]
        res_on = run_on.results[0]

        assert res_on.Ik > res_off.Ik, "Generator in service should increase Ik3"
        assert res_on.ip > res_off.ip, "Generator in service should increase ip"

        # Equivalent positive-sequence impedance should decrease with added source
        assert res_on.Z1.magnitude < res_off.Z1.magnitude

    def test_generator_parameter_mapping_used_by_impedance_model(self):
        gen = SynchronousGenerator(
            id="g_map",
            bus_id="bus_g",
            Sn=20.0,
            Un=10.5,
            Xd_pp=15.0,
            Ra=0.2,
            cos_phi=0.85,
        )

        Z1, _, _ = gen.get_impedance(ref_voltage=10.5)

        zbase = (gen.Un ** 2) / gen.Sn
        expected_x = 0.15 * zbase
        expected_r = 0.002 * zbase

        assert abs(Z1.x - expected_x) < 1e-9, "Xd_pp [%] must map to X1 via percent-to-pu conversion"
        assert abs(Z1.r - expected_r) < 1e-9, "Ra [%] must map to R1 via percent-to-pu conversion"
