"""XLSX export service using openpyxl."""

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models import CalculationRun, Project, NetworkVersion


def generate_calculation_report(
    run: CalculationRun,
    project: Project,
    version: NetworkVersion,
) -> BytesIO:
    """Generate XLSX report for calculation results."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    number_font = Font(name="Consolas")
    number_alignment = Alignment(horizontal="right")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Súhrn"

    # Title
    ws_summary['A1'] = "Protokol o výpočte skratových pomerov"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:F1')

    ws_summary['A2'] = "podľa IEC 60909-0"
    ws_summary['A2'].font = Font(italic=True, color="808080")
    ws_summary.merge_cells('A2:F2')

    # Project info
    info_start = 4
    info_data = [
        ("Projekt:", project.name),
        ("Popis:", project.description or "-"),
        ("Verzia siete:", f"v{version.version_number}"),
        ("Režim výpočtu:", "Maximum (Ik max)" if run.calculation_mode.value == "MAX" else "Minimum (Ik min)"),
        ("Typy porúch:", ", ".join(run.fault_types)),
        ("Dátum výpočtu:", run.completed_at.strftime("%d.%m.%Y %H:%M") if run.completed_at else "-"),
        ("Engine verzia:", run.engine_version),
    ]

    for i, (label, value) in enumerate(info_data):
        ws_summary[f'A{info_start + i}'] = label
        ws_summary[f'A{info_start + i}'].font = Font(bold=True)
        ws_summary[f'B{info_start + i}'] = value

    # Results table
    results_start = info_start + len(info_data) + 2
    ws_summary[f'A{results_start}'] = "Výsledky"
    ws_summary[f'A{results_start}'].font = Font(bold=True, size=12)

    # Group results by bus
    results_by_bus = {}
    for result in run.results:
        if result.bus_id not in results_by_bus:
            results_by_bus[result.bus_id] = {}
        results_by_bus[result.bus_id][result.fault_type.value] = result

    # Header row
    header_row = results_start + 2
    headers = ["Uzol", "Ik3 [kA]", "ip3 [kA]", "Ik2 [kA]", "Ik1 [kA]"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row = header_row + 1
    for bus_id, faults in sorted(results_by_bus.items()):
        ik3 = faults.get('IK3')
        ik2 = faults.get('IK2')
        ik1 = faults.get('IK1')

        ws_summary.cell(row=row, column=1, value=bus_id).border = thin_border

        for col, val in enumerate([
            ik3.Ik if ik3 else None,
            ik3.ip if ik3 else None,
            ik2.Ik if ik2 else None,
            ik1.Ik if ik1 else None,
        ], 2):
            cell = ws_summary.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if val is not None:
                cell.number_format = '0.000'
                cell.font = number_font
                cell.alignment = number_alignment

        row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40
    for col in ['C', 'D', 'E', 'F']:
        ws_summary.column_dimensions[col].width = 12

    # === Sheet 2: Detailed Results ===
    ws_detail = wb.create_sheet("Detaily")

    row = 1
    for bus_id, faults in sorted(results_by_bus.items()):
        # Bus header
        ws_detail.cell(row=row, column=1, value=f"Uzol: {bus_id}")
        ws_detail.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws_detail.merge_cells(f'A{row}:E{row}')
        row += 1

        # Column headers
        col_headers = ["Parameter", "Ik3", "Ik2", "Ik1"]
        for col, header in enumerate(col_headers, 1):
            cell = ws_detail.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1

        ik3 = faults.get('IK3')
        ik2 = faults.get('IK2')
        ik1 = faults.get('IK1')

        # Data rows
        params = [
            ("Ik [kA]", ik3.Ik if ik3 else None, ik2.Ik if ik2 else None, ik1.Ik if ik1 else None),
            ("ip [kA]", ik3.ip if ik3 else None, ik2.ip if ik2 else None, ik1.ip if ik1 else None),
            ("R/X", ik3.R_X_ratio if ik3 else None, ik2.R_X_ratio if ik2 else None, ik1.R_X_ratio if ik1 else None),
            ("c faktor", ik3.c_factor if ik3 else None, ik2.c_factor if ik2 else None, ik1.c_factor if ik1 else None),
            ("R1 [Ω]", ik3.Z1['r'] if ik3 else None, ik2.Z1['r'] if ik2 else None, ik1.Z1['r'] if ik1 else None),
            ("X1 [Ω]", ik3.Z1['x'] if ik3 else None, ik2.Z1['x'] if ik2 else None, ik1.Z1['x'] if ik1 else None),
        ]

        # Add Z0 if present
        if any(f and f.Z0 for f in [ik3, ik2, ik1]):
            params.append((
                "R0 [Ω]",
                ik3.Z0['r'] if ik3 and ik3.Z0 else None,
                ik2.Z0['r'] if ik2 and ik2.Z0 else None,
                ik1.Z0['r'] if ik1 and ik1.Z0 else None,
            ))
            params.append((
                "X0 [Ω]",
                ik3.Z0['x'] if ik3 and ik3.Z0 else None,
                ik2.Z0['x'] if ik2 and ik2.Z0 else None,
                ik1.Z0['x'] if ik1 and ik1.Z0 else None,
            ))

        for param_name, v1, v2, v3 in params:
            ws_detail.cell(row=row, column=1, value=param_name).border = thin_border
            ws_detail.cell(row=row, column=1).font = Font(bold=True)

            for col, val in enumerate([v1, v2, v3], 2):
                cell = ws_detail.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if val is not None:
                    cell.number_format = '0.0000'
                    cell.font = number_font
                    cell.alignment = number_alignment

            row += 1

        row += 1  # Empty row between buses

    # Adjust column widths
    ws_detail.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D']:
        ws_detail.column_dimensions[col].width = 18

    # === Sheet 3: Network Elements ===
    ws_elements = wb.create_sheet("Prvky siete")

    elements = version.elements or {}
    row = 1

    element_types = [
        ("busbars", "Uzly", ["id", "name", "Un", "is_reference"]),
        ("external_grids", "Externé siete", ["id", "name", "bus_id", "Sk_max", "Sk_min", "rx_ratio"]),
        ("lines", "Vedenia", ["id", "name", "bus_from", "bus_to", "length", "r1_per_km", "x1_per_km"]),
        ("transformers_2w", "Transformátory 2W", ["id", "name", "bus_hv", "bus_lv", "Sn", "uk_percent", "vector_group"]),
        ("generators", "Generátory", ["id", "name", "bus_id", "Sn", "Un", "Xd_pp", "cos_phi"]),
        ("motors", "Motory", ["id", "name", "bus_id", "Un", "Pn", "Ia_In"]),
    ]

    for key, title, cols in element_types:
        items = elements.get(key, [])
        if not items:
            continue

        # Section header
        ws_elements.cell(row=row, column=1, value=title)
        ws_elements.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        # Column headers
        for col, header in enumerate(cols, 1):
            cell = ws_elements.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1

        # Data
        for item in items:
            for col, key in enumerate(cols, 1):
                val = item.get(key, "")
                cell = ws_elements.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.font = number_font
                    cell.alignment = number_alignment
            row += 1

        row += 1  # Empty row

    # Adjust column widths
    for col in range(1, 10):
        ws_elements.column_dimensions[get_column_letter(col)].width = 15

    # Save
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
