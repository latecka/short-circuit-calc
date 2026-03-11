"""Network import service - JSON and XLSX."""

import json
import math
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError, field_validator
from pydantic import ConfigDict


# ============================================================================
# Field name mappings for alternative import formats
# ============================================================================

# Element type name mappings
ELEMENT_TYPE_ALIASES = {
    'buses': 'busbars',
    'transformers': 'transformers_2w',
    'trafo_2w': 'transformers_2w',
    'trafo_3w': 'transformers_3w',
}

# Field name mappings per element type
FIELD_ALIASES = {
    'busbars': {
        'un_kv': 'Un',
        'voltage': 'Un',
        'voltage_kv': 'Un',
    },
    'external_grids': {
        'sk_max_mva': 'Sk_max',
        'sk_min_mva': 'Sk_min',
        'sk3max_mva': 'Sk_max',
        'sk3min_mva': 'Sk_min',
        'skmax': 'Sk_max',
        'skmin': 'Sk_min',
        'rx_ratio_max': 'rx_ratio',
    },
    'transformers_2w': {
        'hv_bus_id': 'bus_hv',
        'lv_bus_id': 'bus_lv',
        'bus1': 'bus_hv',
        'bus2': 'bus_lv',
        'sn_mva': 'Sn',
        'rated_power': 'Sn',
        'un1_kv': 'Un_hv',
        'un2_kv': 'Un_lv',
        'un_hv_kv': 'Un_hv',
        'un_lv_kv': 'Un_lv',
        'pkr_kw': 'Pkr',
        'pk_kw': 'Pkr',
        'losses_kw': 'Pkr',
        'uk': 'uk_percent',
    },
    'generators': {
        'sn_mva': 'Sn',
        'rated_power': 'Sn',
        'un_kv': 'Un',
        'rated_voltage': 'Un',
        'xdpp_pu': 'Xd_pp',  # Will be converted from p.u. to %
        'xd_pp_pu': 'Xd_pp',
        'xdpp': 'Xd_pp',
        'ra_pu': 'Ra',  # Will be converted from p.u. to %
        'ra_ohm': 'Ra_ohm',  # Needs special conversion
    },
    'lines': {
        'from_bus': 'bus_from',
        'to_bus': 'bus_to',
        'from_bus_id': 'bus_from',
        'to_bus_id': 'bus_to',
        'length_km': 'length',
        'r1_ohm_per_km': 'r1_per_km',
        'x1_ohm_per_km': 'x1_per_km',
        'r0_ohm_per_km': 'r0_per_km',
        'x0_ohm_per_km': 'x0_per_km',
        'parallel_cables': 'parallel_lines',
    },
    'motors': {
        'un_kv': 'Un',
        'sn_kva': 'Sn_kVA',  # Will be converted to MVA
        'ilr_ratio': 'Ia_In',
        'i_lr_ratio': 'Ia_In',
        'ia_in': 'Ia_In',
    },
}


def normalize_elements(raw: dict[str, Any]) -> dict[str, Any]:
    """
    Normalize element type names and field names to expected format.

    Handles alternative naming conventions from different import sources.
    """
    normalized = {}

    for key, items in raw.items():
        # Skip non-element keys
        if not isinstance(items, list):
            continue

        # Map element type name
        elem_type = ELEMENT_TYPE_ALIASES.get(key, key)

        if elem_type not in normalized:
            normalized[elem_type] = []

        field_map = FIELD_ALIASES.get(elem_type, {})

        for item in items:
            if not isinstance(item, dict):
                continue

            norm_item = {}
            for field, value in item.items():
                # Map field name
                mapped_field = field_map.get(field, field)
                norm_item[mapped_field] = value

            # Special conversions for generators
            if elem_type == 'generators':
                norm_item = _normalize_generator(norm_item)

            # Special conversions for transformers
            if elem_type == 'transformers_2w':
                norm_item = _normalize_transformer_2w(norm_item)

            # Special conversions for motors
            if elem_type == 'motors':
                norm_item = _normalize_motor(norm_item)

            normalized[elem_type].append(norm_item)

    return normalized


def _normalize_generator(item: dict) -> dict:
    """Apply special conversions for generator fields."""
    # Convert Xd_pp from p.u. to % if it's small (< 1.0 suggests p.u.)
    # Note: validation layer also handles this, but we do it here for
    # consistency with ra_ohm conversion
    if 'Xd_pp' in item and item['Xd_pp'] is not None:
        if item['Xd_pp'] < 1.0:
            item['Xd_pp'] = item['Xd_pp'] * 100

    # Convert Ra from ohms to % (special case not handled by validator)
    if 'Ra_ohm' in item and item['Ra_ohm'] is not None:
        # Need Sn and Un to convert
        Sn = item.get('Sn', 0)
        Un = item.get('Un', 0)
        if Sn > 0 and Un > 0:
            Zbase = (Un ** 2) / Sn
            Ra_pu = item['Ra_ohm'] / Zbase
            # Set as p.u. value - the validator will convert to %
            item['Ra'] = Ra_pu
        del item['Ra_ohm']

    # Do NOT convert Ra from p.u. to % here - the validator handles that
    # to avoid double conversion

    return item


def _normalize_motor(item: dict) -> dict:
    """Apply special conversions for motor fields."""
    # Convert Sn from kVA to Pn in kW (assuming cos_phi if available)
    if 'Sn_kVA' in item and item['Sn_kVA'] is not None:
        cos_phi = item.get('cos_phi', 0.85)  # Default power factor
        # Pn = Sn * cos_phi (in same units, so kVA * cos_phi = kW)
        item['Pn'] = item['Sn_kVA'] * cos_phi
        del item['Sn_kVA']

    return item


def _normalize_transformer_2w(item: dict) -> dict:
    """Apply special conversions for 2W transformer fields."""
    # uk might be given without _percent suffix but as percentage
    if 'uk' in item and 'uk_percent' not in item:
        item['uk_percent'] = item['uk']
        del item['uk']

    return item


class ImportError(Exception):
    """Import validation error with details."""
    def __init__(self, message: str, errors: list[dict] = None):
        super().__init__(message)
        self.errors = errors or []


# Validation schemas
class BusbarImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    Un: float = Field(..., gt=0)
    is_reference: bool = False


class ExternalGridImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_id: str
    Sk_max: float = Field(..., gt=0)
    Sk_min: float = Field(..., gt=0)
    rx_ratio: float = Field(0.1, gt=0)
    c_max: float = 1.1
    c_min: float = 1.0
    Z0_Z1_ratio: float = 1.0
    X0_X1_ratio: float = 1.0
    R0_X0_ratio: float | None = None


class LineImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    type: str = "overhead_line"
    bus_from: str
    bus_to: str
    length: float = Field(..., gt=0)
    r1_per_km: float = Field(..., ge=0)
    x1_per_km: float = Field(..., gt=0)
    r0_per_km: float = Field(..., ge=0)
    x0_per_km: float = Field(..., gt=0)
    parallel_lines: int = 1
    in_service: bool = True


class Transformer2WImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_hv: str
    bus_lv: str
    Sn: float = Field(..., gt=0)
    Un_hv: float = Field(..., gt=0)
    Un_lv: float = Field(..., gt=0)
    uk_percent: float = Field(..., gt=0, lt=100)
    Pkr: float = Field(0, ge=0)
    vector_group: str
    tap_position: float = 0.0
    neutral_grounding_hv: str = "isolated"
    neutral_grounding_lv: str = "isolated"
    in_service: bool = True

    @field_validator('vector_group')
    @classmethod
    def validate_vector_group(cls, v):
        valid_groups = ['Dyn11', 'Dyn5', 'Dyn1', 'YNyn0', 'YNd11', 'YNd5', 'Yyn0', 'Dd0', 'Dy11', 'Yd11']
        if v not in valid_groups:
            raise ValueError(f'Invalid vector group. Must be one of: {valid_groups}')
        return v


class Transformer3WImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_hv: str
    bus_mv: str
    bus_lv: str
    Sn_hv: float = Field(..., gt=0)
    Sn_mv: float | None = None
    Sn_lv: float | None = None
    Un_hv: float = Field(..., gt=0)
    Un_mv: float = Field(..., gt=0)
    Un_lv: float = Field(..., gt=0)
    uk_hv_mv_percent: float = Field(..., gt=0)
    uk_hv_lv_percent: float = Field(..., gt=0)
    uk_mv_lv_percent: float = Field(..., gt=0)
    Pkr_hv_mv: float = 0
    Pkr_hv_lv: float = 0
    Pkr_mv_lv: float = 0
    in_service: bool = True


class AutotransformerImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_hv: str
    bus_lv: str
    Sn: float = Field(..., gt=0)
    Un_hv: float = Field(..., gt=0)
    Un_lv: float = Field(..., gt=0)
    uk_percent: float = Field(..., gt=0)
    Pkr: float = 0
    neutral_grounding: str = "grounded"
    has_tertiary_delta: bool = False
    tertiary_Sn: float | None = None
    in_service: bool = True


class GeneratorImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_id: str
    Sn: float = Field(..., gt=0)  # MVA
    Un: float = Field(..., gt=0)  # kV
    Xd_pp: float = Field(..., gt=0)  # % (will auto-convert from p.u. if < 1)
    Ra: float = 0  # % (will auto-convert from p.u. if < 1)
    cos_phi: float = Field(0.85, gt=0, le=1)
    connection: str = "direct"
    in_service: bool = True

    @field_validator('Xd_pp')
    @classmethod
    def convert_xdpp_to_percent(cls, v):
        """Convert Xd'' from p.u. to % if value is < 1 (assumes p.u. input)."""
        # Typical Xd'' values are 10-30%, so if < 1 it's likely in p.u.
        if v < 1.0:
            return v * 100
        return v

    @field_validator('Ra')
    @classmethod
    def convert_ra_to_percent(cls, v):
        """Convert Ra from p.u. to % if value is < 1 (assumes p.u. input)."""
        if v > 0 and v < 1.0:
            return v * 100
        return v


class MotorImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_id: str
    Un: float = Field(..., gt=0)
    input_mode: str = "power"
    Pn: float | None = None
    eta: float | None = None
    cos_phi: float | None = None
    In: float | None = None
    Ia_In: float = Field(..., gt=1)
    pole_pairs: int = 1
    include_in_sc: bool = True
    in_service: bool = True


class PSUImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    generator_id: str
    transformer_id: str
    has_oltc: bool = True
    generator_winding: str | None = None


class ImpedanceImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    bus_from: str
    bus_to: str
    R: float = Field(..., ge=0)
    X: float
    R0: float | None = None
    X0: float | None = None
    in_service: bool = True


class GroundingImpedanceImport(BaseModel):
    model_config = ConfigDict(extra='ignore')

    id: str
    name: str | None = None
    R: float = Field(..., ge=0)
    X: float = Field(..., ge=0)


ELEMENT_SCHEMAS = {
    'busbars': BusbarImport,
    'external_grids': ExternalGridImport,
    'lines': LineImport,
    'transformers_2w': Transformer2WImport,
    'transformers_3w': Transformer3WImport,
    'autotransformers': AutotransformerImport,
    'generators': GeneratorImport,
    'motors': MotorImport,
    'psus': PSUImport,
    'impedances': ImpedanceImport,
    'grounding_impedances': GroundingImpedanceImport,
}


def validate_elements(elements: dict[str, Any]) -> tuple[dict[str, Any], list[dict]]:
    """
    Validate imported elements.

    Returns:
        Tuple of (validated_elements, errors)
    """
    validated = {}
    errors = []
    all_bus_ids = set()
    all_element_ids = set()

    # First pass: collect all bus IDs
    for bus in elements.get('busbars', []):
        if isinstance(bus, dict) and 'id' in bus:
            all_bus_ids.add(bus['id'])

    # Validate each element type
    for elem_type, schema in ELEMENT_SCHEMAS.items():
        items = elements.get(elem_type, [])
        validated[elem_type] = []

        for i, item in enumerate(items):
            try:
                # Validate with Pydantic
                validated_item = schema(**item)
                item_dict = validated_item.model_dump()

                # Check for duplicate IDs
                if item_dict['id'] in all_element_ids:
                    errors.append({
                        'type': elem_type,
                        'index': i,
                        'id': item_dict['id'],
                        'error': f"Duplicate element ID: {item_dict['id']}",
                    })
                    continue

                all_element_ids.add(item_dict['id'])

                # Check bus references
                bus_fields = ['bus_id', 'bus_from', 'bus_to', 'bus_hv', 'bus_lv', 'bus_mv']
                for field in bus_fields:
                    if field in item_dict and item_dict[field]:
                        if item_dict[field] not in all_bus_ids:
                            errors.append({
                                'type': elem_type,
                                'index': i,
                                'id': item_dict.get('id'),
                                'error': f"Referenced bus '{item_dict[field]}' does not exist",
                            })

                validated[elem_type].append(item_dict)

            except ValidationError as e:
                for err in e.errors():
                    errors.append({
                        'type': elem_type,
                        'index': i,
                        'id': item.get('id', 'unknown'),
                        'field': '.'.join(str(x) for x in err['loc']),
                        'error': err['msg'],
                    })

    # Check generator/transformer references in PSUs
    gen_ids = {g['id'] for g in validated.get('generators', [])}
    tr_ids = {t['id'] for t in validated.get('transformers_2w', [])}
    tr_ids.update(t['id'] for t in validated.get('transformers_3w', []))

    for i, psu in enumerate(validated.get('psus', [])):
        if psu['generator_id'] not in gen_ids:
            errors.append({
                'type': 'psus',
                'index': i,
                'id': psu['id'],
                'error': f"Referenced generator '{psu['generator_id']}' does not exist",
            })
        if psu['transformer_id'] not in tr_ids:
            errors.append({
                'type': 'psus',
                'index': i,
                'id': psu['id'],
                'error': f"Referenced transformer '{psu['transformer_id']}' does not exist",
            })

    # Build bus voltage lookup for topology validation
    bus_voltages = {bus['id']: bus['Un'] for bus in validated.get('busbars', [])}

    # Validate lines: cannot connect buses with different voltage levels
    for i, line in enumerate(validated.get('lines', [])):
        bus_from = line.get('bus_from')
        bus_to = line.get('bus_to')

        if bus_from in bus_voltages and bus_to in bus_voltages:
            un_from = bus_voltages[bus_from]
            un_to = bus_voltages[bus_to]

            if un_from != un_to:
                errors.append({
                    'type': 'lines',
                    'index': i,
                    'id': line['id'],
                    'error': (
                        f"Vedenie '{line['id']}' spája uzly s rôznymi napäťovými hladinami: "
                        f"{bus_from} ({un_from} kV) → {bus_to} ({un_to} kV). "
                        f"Vedenia môžu spájať len uzly s rovnakým Un."
                    ),
                })

    return validated, errors


def import_from_json(data: bytes | str) -> dict[str, Any]:
    """
    Import network elements from JSON.

    Supports two formats:
    1. Direct elements: {"busbars": [...], "lines": [...], ...}
    2. Export format: {"export_version": "1.0", "network_elements": {...}, ...}

    Args:
        data: JSON string or bytes

    Returns:
        Validated elements dictionary

    Raises:
        ImportError: If validation fails
    """
    try:
        if isinstance(data, bytes):
            data = data.decode('utf-8')
        raw_data = json.loads(data)
    except json.JSONDecodeError as e:
        raise ImportError(f"Invalid JSON: {e}")

    if not isinstance(raw_data, dict):
        raise ImportError("JSON must be an object with element arrays")

    # Check if this is export format (has network_elements wrapper)
    if 'network_elements' in raw_data and isinstance(raw_data['network_elements'], dict):
        raw_elements = raw_data['network_elements']
    else:
        # Direct elements format
        raw_elements = raw_data

    # Normalize field names from alternative formats
    normalized = normalize_elements(raw_elements)

    validated, errors = validate_elements(normalized)

    if errors:
        raise ImportError(
            f"Validation failed with {len(errors)} error(s)",
            errors=errors
        )

    return validated


def import_from_xlsx(data: bytes) -> dict[str, Any]:
    """
    Import network elements from XLSX.

    Expected format:
    - Each sheet corresponds to an element type
    - First row contains column headers matching field names
    - Subsequent rows contain element data

    Args:
        data: XLSX file bytes

    Returns:
        Validated elements dictionary

    Raises:
        ImportError: If validation fails
    """
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ImportError(f"Invalid XLSX file: {e}")

    # Sheet name to element type mapping
    sheet_mapping = {
        'Uzly': 'busbars',
        'Busbars': 'busbars',
        'busbars': 'busbars',
        'Externé siete': 'external_grids',
        'External Grids': 'external_grids',
        'external_grids': 'external_grids',
        'Vedenia': 'lines',
        'Lines': 'lines',
        'lines': 'lines',
        'Transformátory 2W': 'transformers_2w',
        'Transformers 2W': 'transformers_2w',
        'transformers_2w': 'transformers_2w',
        'Transformátory 3W': 'transformers_3w',
        'Transformers 3W': 'transformers_3w',
        'transformers_3w': 'transformers_3w',
        'Autotransformátory': 'autotransformers',
        'Autotransformers': 'autotransformers',
        'autotransformers': 'autotransformers',
        'Generátory': 'generators',
        'Generators': 'generators',
        'generators': 'generators',
        'Motory': 'motors',
        'Motors': 'motors',
        'motors': 'motors',
        'PSU': 'psus',
        'psus': 'psus',
        'Impedancie': 'impedances',
        'Impedances': 'impedances',
        'impedances': 'impedances',
        'Zemniace impedancie': 'grounding_impedances',
        'Grounding Impedances': 'grounding_impedances',
        'grounding_impedances': 'grounding_impedances',
    }

    raw_elements = {key: [] for key in ELEMENT_SCHEMAS.keys()}

    for sheet_name in wb.sheetnames:
        elem_type = sheet_mapping.get(sheet_name)
        if not elem_type:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # First row is header
        headers = [str(h).strip() if h else '' for h in rows[0]]

        # Parse data rows
        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):  # Skip empty rows
                continue

            item = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    header = headers[col_idx]
                    # Convert Excel values
                    if value is not None:
                        # Handle boolean
                        if isinstance(value, bool):
                            item[header] = value
                        elif isinstance(value, str):
                            lower_val = value.lower().strip()
                            if lower_val in ('true', 'áno', 'yes', '1'):
                                item[header] = True
                            elif lower_val in ('false', 'nie', 'no', '0'):
                                item[header] = False
                            else:
                                item[header] = value.strip()
                        else:
                            item[header] = value

            if item.get('id'):  # Only add if has ID
                raw_elements[elem_type].append(item)

    wb.close()

    validated, errors = validate_elements(raw_elements)

    if errors:
        raise ImportError(
            f"Validation failed with {len(errors)} error(s)",
            errors=errors
        )

    return validated


def generate_template_xlsx() -> bytes:
    """Generate an empty XLSX template for import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # Sheet definitions with headers
    sheets = [
        ('Busbars', ['id', 'name', 'Un', 'is_reference']),
        ('External Grids', ['id', 'name', 'bus_id', 'Sk_max', 'Sk_min', 'rx_ratio', 'c_max', 'c_min']),
        ('Lines', ['id', 'name', 'type', 'bus_from', 'bus_to', 'length', 'r1_per_km', 'x1_per_km', 'r0_per_km', 'x0_per_km']),
        ('Transformers 2W', ['id', 'name', 'bus_hv', 'bus_lv', 'Sn', 'Un_hv', 'Un_lv', 'uk_percent', 'Pkr', 'vector_group']),
        ('Transformers 3W', ['id', 'name', 'bus_hv', 'bus_mv', 'bus_lv', 'Sn_hv', 'Un_hv', 'Un_mv', 'Un_lv', 'uk_hv_mv_percent', 'uk_hv_lv_percent', 'uk_mv_lv_percent']),
        ('Autotransformers', ['id', 'name', 'bus_hv', 'bus_lv', 'Sn', 'Un_hv', 'Un_lv', 'uk_percent', 'neutral_grounding', 'has_tertiary_delta']),
        ('Generators', ['id', 'name', 'bus_id', 'Sn', 'Un', 'Xd_pp', 'Ra', 'cos_phi']),
        ('Motors', ['id', 'name', 'bus_id', 'Un', 'Pn', 'eta', 'cos_phi', 'Ia_In', 'pole_pairs']),
        ('PSU', ['id', 'name', 'generator_id', 'transformer_id', 'has_oltc']),
        ('Impedances', ['id', 'name', 'bus_from', 'bus_to', 'R', 'X', 'R0', 'X0']),
        ('Grounding Impedances', ['id', 'name', 'R', 'X']),
    ]

    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, headers in sheets:
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + col)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
